import openpyxl
from openpyxl.utils import get_column_letter
import sys

# Set UTF-8 encoding for Windows console
sys.stdout.reconfigure(encoding='utf-8')

# Load workbook
wb = openpyxl.load_workbook('docs/Облік нальоту.xlsx', data_only=True)

print("Available sheets:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Try to find the "Підсумки" sheet
target_sheet = None
for sheet_name in wb.sheetnames:
    if 'підсумки' in sheet_name.lower() or 'пiдсумки' in sheet_name.lower():
        target_sheet = wb[sheet_name]
        print(f"\nFound sheet: {sheet_name}")
        break

if not target_sheet:
    # Try the third sheet (based on the garbled output showing it might be third)
    if len(wb.sheetnames) >= 3:
        target_sheet = wb[wb.sheetnames[2]]
        print(f"\nUsing third sheet: {wb.sheetnames[2]}")

if target_sheet:
    print(f"\n{'='*80}")
    print(f"SHEET STRUCTURE: {target_sheet.title}")
    print(f"{'='*80}")

    # Get dimensions
    max_row = target_sheet.max_row
    max_col = target_sheet.max_column
    print(f"\nDimensions: {max_row} rows × {max_col} columns")

    # Print header rows (first 5 rows to capture any multi-row headers)
    print(f"\n{'='*80}")
    print("HEADERS (first 10 rows):")
    print(f"{'='*80}")

    for row_idx in range(1, min(11, max_row + 1)):
        row_values = []
        for col_idx in range(1, max_col + 1):
            cell = target_sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                value = ""
            row_values.append(str(value))

        print(f"Row {row_idx:2d}: {' | '.join(row_values)}")

    # Print column headers mapping
    print(f"\n{'='*80}")
    print("COLUMN MAPPING:")
    print(f"{'='*80}")

    # Assuming headers are in row 1, but let's check first few rows
    for check_row in range(1, min(6, max_row + 1)):
        has_headers = False
        for col_idx in range(1, max_col + 1):
            cell_value = target_sheet.cell(row=check_row, column=col_idx).value
            if cell_value and isinstance(cell_value, str) and len(cell_value) > 0:
                has_headers = True
                break

        if has_headers:
            print(f"\nHeaders in row {check_row}:")
            for col_idx in range(1, max_col + 1):
                cell = target_sheet.cell(row=check_row, column=col_idx)
                col_letter = get_column_letter(col_idx)
                value = cell.value if cell.value is not None else ""
                print(f"  {col_letter} (col {col_idx:2d}): {value}")

    # Print sample data rows
    print(f"\n{'='*80}")
    print("SAMPLE DATA (rows 11-20):")
    print(f"{'='*80}")

    for row_idx in range(11, min(21, max_row + 1)):
        row_values = []
        for col_idx in range(1, min(max_col + 1, 15)):  # Limit to first 15 columns for readability
            cell = target_sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                value = ""
            row_values.append(str(value)[:20])  # Truncate long values

        if any(val.strip() for val in row_values):  # Only print non-empty rows
            print(f"Row {row_idx:2d}: {' | '.join(row_values)}")

    # Print merged cells info
    print(f"\n{'='*80}")
    print("MERGED CELLS:")
    print(f"{'='*80}")
    if hasattr(target_sheet, 'merged_cells') and target_sheet.merged_cells:
        for merged_range in list(target_sheet.merged_cells.ranges)[:20]:  # First 20 merged ranges
            print(f"  {merged_range}")
    else:
        print("  No merged cells found")

    # Check for any formulas or special formatting
    print(f"\n{'='*80}")
    print("CELLS WITH FORMULAS (sample):")
    print(f"{'='*80}")
    formula_count = 0
    for row_idx in range(1, min(50, max_row + 1)):
        for col_idx in range(1, max_col + 1):
            cell = target_sheet.cell(row=row_idx, column=col_idx)
            if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"  {get_column_letter(col_idx)}{row_idx}: {cell.value}")
                formula_count += 1
                if formula_count >= 10:
                    break
        if formula_count >= 10:
            break

    if formula_count == 0:
        print("  No formulas found in first 50 rows")

else:
    print("\nCould not find Підсумки sheet")
