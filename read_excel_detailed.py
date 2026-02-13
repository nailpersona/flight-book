import openpyxl
from openpyxl.utils import get_column_letter
import sys

# Set UTF-8 encoding for Windows console
sys.stdout.reconfigure(encoding='utf-8')

# Load workbook
wb = openpyxl.load_workbook('docs/Облік нальоту.xlsx', data_only=True)

target_sheet = wb['Підсумки']

print("="*100)
print("ПІДСУМКИ SHEET - DETAILED STRUCTURE")
print("="*100)

# Get all row 4, 5, 6 headers to understand the multi-level header structure
print("\nMULTI-LEVEL HEADERS:")
print("="*100)

max_col = target_sheet.max_column

for col_idx in range(1, max_col + 1):
    col_letter = get_column_letter(col_idx)
    row4 = target_sheet.cell(row=4, column=col_idx).value or ""
    row5 = target_sheet.cell(row=5, column=col_idx).value or ""
    row6 = target_sheet.cell(row=6, column=col_idx).value or ""

    print(f"{col_letter:3s} (col {col_idx:2d}): Row4=[{str(row4)[:40]}] Row5=[{str(row5)[:40]}] Row6=[{str(row6)[:40]}]")

print("\n" + "="*100)
print("ALL DATA ROWS (complete):")
print("="*100)

max_row = target_sheet.max_row

for row_idx in range(7, max_row + 1):
    row_data = []
    has_data = False

    for col_idx in range(1, max_col + 1):
        cell = target_sheet.cell(row=row_idx, column=col_idx)
        value = cell.value
        if value is not None and str(value).strip() != "":
            has_data = True
        row_data.append(str(value) if value is not None else "")

    if has_data:
        print(f"\nRow {row_idx:2d}:")
        for col_idx, value in enumerate(row_data, start=1):
            if value:  # Only print non-empty cells
                col_letter = get_column_letter(col_idx)
                print(f"  {col_letter:3s}: {value}")

# Check merged cells to understand header groupings
print("\n" + "="*100)
print("MERGED CELLS (detailed):")
print("="*100)
if hasattr(target_sheet, 'merged_cells') and target_sheet.merged_cells:
    for merged_range in list(target_sheet.merged_cells.ranges):
        # Get the value from the top-left cell of the merged range
        top_left_cell = target_sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
        value = top_left_cell.value or ""
        print(f"  {merged_range} = [{value}]")

# Analyze column groups based on headers
print("\n" + "="*100)
print("INTERPRETED COLUMN STRUCTURE:")
print("="*100)

print("\nMain groups based on Row 4 headers:")
print("  A-B: Row identifiers (№, Тип ПС)")
print("  C-E: В період (current period)")
print("  F-H: З початку 2025 року (since start of year)")
print("  I-L: Випроб. польоти (test flights including МЛВ)")
print("  M+: Additional columns (if any)")

print("\nWithin 'В період' (C-E):")
row5_c = target_sheet.cell(row=5, column=3).value
row5_d = target_sheet.cell(row=5, column=4).value
row5_e = target_sheet.cell(row=5, column=5).value
print(f"  C: {row5_c}")
print(f"  D: {row5_d}")
print(f"  E: {row5_e}")

print("\nWithin 'З початку 2025 року' (F-H):")
row5_f = target_sheet.cell(row=5, column=6).value
row5_g = target_sheet.cell(row=5, column=7).value
row5_h = target_sheet.cell(row=5, column=8).value
print(f"  F: {row5_f}")
print(f"  G: {row5_g}")
print(f"  H: {row5_h}")

print("\nWithin 'Випроб. польоти (з них МЛВ)' (I-L):")
row6_i = target_sheet.cell(row=6, column=9).value
row6_j = target_sheet.cell(row=6, column=10).value
row6_k = target_sheet.cell(row=6, column=11).value
row6_l = target_sheet.cell(row=6, column=12).value
print(f"  I: {row6_i}")
print(f"  J: {row6_j}")
print(f"  K: {row6_k}")
print(f"  L: {row6_l}")
